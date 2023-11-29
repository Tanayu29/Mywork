import logging
import os
import tkinter as tk
import openpyxl
import requests
from tkinter import filedialog
from datetime import datetime

# ログの設定を行う関数（前の例と同じ）
def setup_logging():
    current_date = datetime.now().strftime("%Y-%m-%d")
    log_filename = f"{current_date}.log"
    log_directory = os.path.join(os.getcwd(), 'log')
    if not os.path.exists(log_directory):
        os.makedirs(log_directory)
    log_file_path = os.path.join(log_directory, log_filename)
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s:%(levelname)s:%(message)s',
        handlers=[
            logging.FileHandler(log_file_path, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

# ウィンドウが閉じられるときに実行される関数
def on_close():
    logging.info("ログ終了")
    root.destroy()
    
# ログ設定の初期化
setup_logging()    

# ログの開始メッセージ
logging.info("ログ開始")


def update_user_in_crowd(user_info, api_url, api_user, api_password):
    # Crowd REST APIのユーザー更新エンドポイント
    url = f'{api_url}/rest/usermanagement/1/user/{user_info["username"]}'

    # リクエストボディの作成
    data = {
        "name": user_info['username'],
        "email": user_info['email'],
        "fullname": user_info['fullname'],
        # その他必要な情報（パスワードなど）
    }

    # ベーシック認証を使用してリクエストを送信
    try:
        response = requests.put(url, json=data, auth=(api_user, api_password))

        # レスポンスのチェック
        if response.status_code == 200:
            logging.info(f"ユーザー {user_info['username']} の更新に成功しました。")
        else:
            logging.error(f"ユーザー更新に失敗しました: {response.text}")
            return False
    except requests.RequestException as e:
        logging.error(f"リクエスト中にエラーが発生しました: {e}")
        return False

    return True

def add_user_to_crowd(user_info_list):
    for user_info in user_info_list:
        # 必要なユーザー情報が存在し、かつ値が空でないか確認
        if 'username' in user_info and user_info['username'] and \
        'email' in user_info and user_info['email'] and \
        'fullname' in user_info and user_info['fullname']:
            # Crowd REST APIのエンドポイント
            url = 'https://your-crowd-api-endpoint/usermanagement/1/user'

            # リクエストボディの作成
            data = {
                "name": user_info['username'],
                "email": user_info['email'],
                "fullname": user_info['fullname'],
                # その他必要な情報（パスワードなど）
            }

            # API認証情報（ベーシック認証）
            api_user = 'your-api-username'
            api_password = 'your-api-password'

            # ベーシック認証を使用してリクエストを送信
            try:
                response = requests.post(url, json=data, auth=(api_user, api_password))

                # レスポンスのチェック
                if response.status_code == 201:
                    logging.info(f"ユーザー {user_info['username']} の追加に成功しました。")
                elif response.status_code == 400:
                    logging.info(f"ユーザー {user_info['username']} は既に存在します。ユーザー情報を更新します。")
                    # ユーザー更新のAPIを呼び出す
                    if not update_user_in_crowd(user_info, url, api_user, api_password):
                        return False
                else:
                    logging.error(f"ユーザー追加に失敗しました: {response.text}")
                    return False
            except requests.RequestException as e:
                logging.error(f"リクエスト中にエラーが発生しました: {e}")
                return False
        else:
            logging.error("必要なユーザー情報が不足しています。")
            return False

    return True



def open_file_dialog():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        # Excelファイルを開く
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        # ヘッダー行を検索し、列番号を取得
        col_server = col_username = col_firstname = col_lastname = col_email = None
        for row in sheet.iter_rows(values_only=True):
            if '登録サーバー' in row and 'UserName' in row and 'FirstName' in row and 'LastName' in row and 'e-mail' in row:
                col_server = row.index('登録サーバー') + 1
                col_username = row.index('UserName') + 1
                col_firstname = row.index('FirstName') + 1
                col_lastname = row.index('LastName') + 1
                col_email = row.index('e-mail') + 1
                break

        if None in [col_server, col_username, col_firstname, col_lastname, col_email]:
            logging.error("ヘッダー行が見つかりません。")
            return
        
        # シートの最終行までユーザー情報をチェックし、リストに格納
        for row in sheet.iter_rows(min_row=2, values_only=True):
            user_info_list = {
                'server': row[col_server - 1],
                'username': row[col_username - 1],
                'firstname': row[col_firstname - 1],
                'lastname': row[col_lastname - 1],
                'email': row[col_email - 1],
                'fullname': f"{row[col_firstname - 1]} {row[col_lastname - 1]}"
            }
            
            # ユーザー情報をログ出力
            logging.info(f"ユーザー情報: {user_info_list}")
            
            # Crowd REST APIの「ユーザー追加」を実行
            try:
                add_user_to_crowd(user_info_list)
            except Exception as e:
                logging.error(f"ユーザー追加中にエラーが発生しました: {e}")
                break  # エラーが発生した場合は処理を中断

# メインウィンドウを作成
root = tk.Tk()
root.title("ユーザー情報管理")

# 上部のボタンの作成
frame_buttons = tk.Frame(root)
frame_buttons.grid(row=0, column=0, columnspan=2, pady=10)

# "ユーザー情報変更（申請書読み込み）"ボタンを作成
button_file = tk.Button(root, text="ユーザー情報変更（申請書読み込み）", command=open_file_dialog)
button_file.grid(in_=frame_buttons, row=0, column=0, pady=10)

# 終了プロトコルの設定
root.protocol("WM_DELETE_WINDOW", on_close)

# GUIを起動
root.mainloop()